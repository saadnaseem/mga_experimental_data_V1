"""Clean the 48-well plate OD600 kinetics for the top-5 media follow-up
(DBTL round 1).

Reads the instrument xlsx in the same folder as this script and writes:
    plate_map.csv
    omics_growth_long.csv
    omics_growth_by_condition.csv
    omics_growth_summary_per_condition.csv

Run:
    python clean_omics.py

Dependencies: numpy, pandas, scipy, openpyxl (see ../requirements.txt).
"""
from pathlib import Path
import datetime

import numpy as np
import pandas as pd
import openpyxl
from scipy.signal import savgol_filter

HERE = Path(__file__).resolve().parent
XLSX = HERE / '05222026_top5_media_dbtl1.xlsx'

# Plate map (source of truth: the 'well designs' sheet inside the xlsx).
PLATE_MAP = {
    'MPOB_008':   ['B2', 'B3', 'B4'],   # top-5 design 1
    'MPOB_058':   ['B5', 'B6', 'B7'],   # top-5 design 2
    'MPOB_045':   ['C2', 'C3', 'C4'],   # top-5 design 3
    'MPOB_041':   ['C5', 'C6', 'C7'],   # top-5 design 4
    'MPOB_068':   ['D2', 'D3', 'D4'],   # top-5 design 5
    'ctrl_media': ['D5', 'D6', 'D7'],   # in-plate control (same as the SN1 Biolog control)
    'MP':         ['E2', 'E3', 'E4'],   # standard M. extorquens minimal medium reference
}
BLANK_WELLS = ['B1', 'C1', 'D1', 'E1', 'B8', 'C8', 'D8', 'E8']
UNUSED_WELLS = ([f'{r}{c}' for r in 'AF' for c in range(1, 9)]
                + [f'E{c}' for c in (5, 6, 7)])


def to_hours(v):
    """Convert Excel time value to hours since assay start."""
    if isinstance(v, datetime.timedelta):
        return v.total_seconds() / 3600.0
    if isinstance(v, datetime.time):
        return (v.hour * 3600 + v.minute * 60 + v.second) / 3600.0
    return np.nan


def mu_max(t, y, win_h=4.0, min_r2=0.95, min_delta=0.015,
           min_pts=6, sw=7, so=2, eps=1e-4):
    """Specific growth rate (h^-1) via Savitzky–Golay-smoothed, baseline-
    subtracted absorbance and a sliding 4-h log-linear fit. NaN if no window
    satisfies (Δabs ≥ 0.015 everywhere, ≥ 6 points, R² ≥ 0.95).
    """
    order = np.argsort(t)
    t = np.asarray(t, dtype=float)[order]
    y = np.asarray(y, dtype=float)[order]
    if len(t) < max(sw, 4):
        return np.nan
    dy = y - y[0]
    win = min(sw, len(dy) - (1 - len(dy) % 2))
    if win < so + 2:
        dys = dy
    else:
        if win % 2 == 0:
            win -= 1
        dys = savgol_filter(dy, win, so)
    if dys.max() < min_delta:
        return np.nan
    ly = np.log(np.clip(dys, eps, None))
    best = -np.inf
    for i in range(len(t)):
        m = (t >= t[i]) & (t <= t[i] + win_h)
        if m.sum() < min_pts:
            continue
        if dys[m].min() < min_delta:
            continue
        tt, yy = t[m], ly[m]
        s, b = np.polyfit(tt, yy, 1)
        yf = s * tt + b
        ss_res = float(((yy - yf) ** 2).sum())
        ss_tot = float(((yy - yy.mean()) ** 2).sum())
        if ss_tot < 1e-12:
            continue
        r2 = 1.0 - ss_res / ss_tot
        if r2 < min_r2 or s <= 0:
            continue
        if s > best:
            best = float(s)
    return best if np.isfinite(best) else np.nan


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Sheet1']
    header = [c.value for c in ws[1]]
    col_of = {name: i for i, name in enumerate(header) if name}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    t_h = np.array([to_hours(r[col_of['Time']]) for r in rows])
    temp = np.array([r[col_of['T° abs600:600']] for r in rows], dtype=float)
    wells_present = [c for c in header if c and c not in ('Time', 'T° abs600:600')]
    od = pd.DataFrame({w: [r[col_of[w]] for r in rows] for w in wells_present},
                      index=t_h, dtype=float)
    od.index.name = 'time_h'

    # ---- plate map table ----
    map_rows = []
    for cond, wells in PLATE_MAP.items():
        for i, w in enumerate(wells, 1):
            map_rows.append({'well': w, 'row': w[0], 'col': int(w[1:]),
                             'condition': cond, 'replicate': i, 'role': 'sample'})
    for w in BLANK_WELLS:
        map_rows.append({'well': w, 'row': w[0], 'col': int(w[1:]),
                         'condition': 'blank', 'replicate': np.nan,
                         'role': 'blank (media only, no cells)'})
    for w in UNUSED_WELLS:
        map_rows.append({'well': w, 'row': w[0], 'col': int(w[1:]),
                         'condition': 'unused', 'replicate': np.nan, 'role': 'unused'})
    plate_map = pd.DataFrame(map_rows).sort_values(['row', 'col']).reset_index(drop=True)
    plate_map.to_csv(HERE / 'plate_map.csv', index=False)

    # ---- blank series and long-format cleaned kinetics ----
    blank_series = od[BLANK_WELLS].mean(axis=1)
    long_rows = []
    for cond, wells in PLATE_MAP.items():
        for rep, w in enumerate(wells, 1):
            for t, raw, blk, tm in zip(od.index, od[w], blank_series, temp):
                long_rows.append({
                    'time_h': round(float(t), 4),
                    'well': w, 'row': w[0], 'col': int(w[1:]),
                    'condition': cond, 'replicate': rep,
                    'OD600_raw': float(raw),
                    'blank_mean': round(float(blk), 4),
                    'OD600_blank_subtracted': round(float(raw) - float(blk), 4),
                    'temperature_C': float(tm),
                })
    long = pd.DataFrame(long_rows)
    long.to_csv(HERE / 'omics_growth_long.csv', index=False)

    # ---- mean ± std per condition per timepoint (blank-subtracted) ----
    by_cond = (long.groupby(['condition', 'time_h'])['OD600_blank_subtracted']
                    .agg(['mean', 'std', 'count'])
                    .reset_index()
                    .rename(columns={'mean': 'OD600_bsub_mean',
                                     'std':  'OD600_bsub_std',
                                     'count': 'n_reps'}))
    by_cond['OD600_bsub_mean'] = by_cond['OD600_bsub_mean'].round(4)
    by_cond['OD600_bsub_std']  = by_cond['OD600_bsub_std'].round(4)
    by_cond.to_csv(HERE / 'omics_growth_by_condition.csv', index=False)

    # ---- per-condition kinetic summary ----
    summary_rows = []
    for cond, wells in PLATE_MAP.items():
        per_well = []
        for w in wells:
            y_raw = od[w].to_numpy()
            y_bs = y_raw - blank_series.to_numpy()
            per_well.append({
                'OD_final': float(y_bs[-1]),
                'OD_max':   float(y_bs.max()),
                'mu_max':   mu_max(od.index.to_numpy(), y_raw),
            })
        df = pd.DataFrame(per_well)
        m = df.mean(); s = df.std(ddof=1)
        mu = m['mu_max']
        summary_rows.append({
            'condition': cond,
            'n_reps': len(wells),
            'wells': ', '.join(wells),
            'OD_final_mean':      round(m['OD_final'], 3),
            'OD_final_std':       round(s['OD_final'], 3),
            'OD_max_mean':        round(m['OD_max'], 3),
            'OD_max_std':         round(s['OD_max'], 3),
            'mu_max_mean_per_h':  round(mu, 3) if np.isfinite(mu) else np.nan,
            'mu_max_std_per_h':   round(s['mu_max'], 3) if np.isfinite(s['mu_max']) else np.nan,
            'doubling_time_h':    round(float(np.log(2) / mu), 2) if np.isfinite(mu) and mu > 0 else np.nan,
        })
    summary = pd.DataFrame(summary_rows)
    summary.to_csv(HERE / 'omics_growth_summary_per_condition.csv', index=False)

    print(f'{len(od)} timepoints × {len(wells_present)} wells loaded from {XLSX.name}')
    print('wrote:')
    for f in ('plate_map.csv', 'omics_growth_long.csv',
              'omics_growth_by_condition.csv',
              'omics_growth_summary_per_condition.csv'):
        print(f'  {HERE / f}')
    print()
    print(summary.to_string(index=False))


if __name__ == '__main__':
    main()
